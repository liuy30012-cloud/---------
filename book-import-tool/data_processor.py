"""Excel 批量导出工具"""
import json
import os
from pathlib import Path
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import config


class ExcelExporter:
    """Excel 导出器"""

    def __init__(self):
        self.failed_records = []
        self._ensure_directories()

    def _ensure_directories(self):
        """确保必要的目录存在"""
        os.makedirs(config.PROCESSED_DATA_DIR, exist_ok=True)

    def export_all(self):
        """导出所有批次数据到 Excel"""
        print("开始导出 Excel 文件...")

        # 获取所有 batch_*.json 文件
        raw_dir = Path(config.RAW_DATA_DIR)
        batch_files = sorted(raw_dir.glob("batch_*.json"))

        if not batch_files:
            print(f"未找到批次文件在 {config.RAW_DATA_DIR}")
            return

        print(f"找到 {len(batch_files)} 个批次文件")

        # 读取所有图书数据
        all_books = []
        for batch_file in batch_files:
            try:
                with open(batch_file, 'r', encoding='utf-8') as f:
                    batch_data = json.load(f)
                    books = batch_data.get('books', [])
                    all_books.extend(books)
                    print(f"已加载 {batch_file.name}: {len(books)} 本图书")
            except Exception as e:
                print(f"读取 {batch_file.name} 失败: {e}")

        print(f"总共加载 {len(all_books)} 本图书")

        # 按批次大小分割并导出
        total_files = (len(all_books) + config.EXCEL_BATCH_SIZE - 1) // config.EXCEL_BATCH_SIZE

        for i in range(0, len(all_books), config.EXCEL_BATCH_SIZE):
            batch_books = all_books[i:i + config.EXCEL_BATCH_SIZE]
            file_num = i // config.EXCEL_BATCH_SIZE + 1

            output_file = os.path.join(
                config.PROCESSED_DATA_DIR,
                f"books_{file_num:03d}.xlsx"
            )

            print(f"正在导出 {output_file} ({len(batch_books)} 本图书)...")
            self._export_batch(batch_books, output_file)
            print(f"已完成 {file_num}/{total_files}")

        # 保存失败记录
        if self.failed_records:
            self._save_failed_records()
            print(f"有 {len(self.failed_records)} 条记录处理失败，已保存到 {config.FAILED_FILE}")

        print("导出完成！")

    def _export_batch(self, books: List[Dict[str, Any]], output_file: str):
        """导出一批图书到 Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "图书数据"

        # 定义表头
        headers = [
            "ISBN", "书名", "作者", "出版社", "出版日期", "价格",
            "页数", "装帧", "分类", "简介", "封面URL", "豆瓣评分",
            "位置", "状态", "可用性", "流通策略", "总副本数", "语言"
        ]

        # 写入表头
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 写入数据
        for row_num, book in enumerate(books, 2):
            try:
                row_data = self._extract_book_data(book)
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            except Exception as e:
                print(f"处理图书数据失败: {e}")
                self.failed_records.append({
                    "book": book,
                    "error": str(e)
                })

        # 调整列宽
        column_widths = {
            'A': 15,  # ISBN
            'B': 30,  # 书名
            'C': 20,  # 作者
            'D': 20,  # 出版社
            'E': 12,  # 出版日期
            'F': 10,  # 价格
            'G': 8,   # 页数
            'H': 10,  # 装帧
            'I': 15,  # 分类
            'J': 50,  # 简介
            'K': 40,  # 封面URL
            'L': 10,  # 豆瓣评分
            'M': 12,  # 位置
            'N': 12,  # 状态
            'O': 12,  # 可用性
            'P': 12,  # 流通策略
            'Q': 10,  # 总副本数
            'R': 8    # 语言
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # 冻结首行
        ws.freeze_panes = "A2"

        # 保存文件
        wb.save(output_file)

    def _extract_book_data(self, book: Dict[str, Any]) -> List[Any]:
        """从图书对象中提取数据"""
        # ISBN
        isbn = book.get('isbn13', '') or book.get('isbn10', '')

        # 书名
        title = book.get('title', '')

        # 作者（合并为字符串）
        authors = book.get('author', [])
        author_str = ', '.join(authors) if isinstance(authors, list) else str(authors)

        # 出版社
        publisher = book.get('publisher', '')

        # 出版日期
        pubdate = book.get('pubdate', '')

        # 价格
        price = book.get('price', '')

        # 页数
        pages = book.get('pages', '')

        # 装帧
        binding = book.get('binding', '')

        # 分类（取第一个标签）
        tags = book.get('tags', [])
        category = tags[0].get('name', '') if tags else ''

        # 简介
        summary = book.get('summary', '')

        # 封面URL
        images = book.get('images', {})
        cover_url = images.get('large', '') or images.get('medium', '') or images.get('small', '')

        # 豆瓣评分
        rating = book.get('rating', {})
        rating_average = rating.get('average', '')

        # 默认值
        location = config.DEFAULT_LOCATION
        status = config.DEFAULT_STATUS
        availability = config.DEFAULT_AVAILABILITY
        circulation_policy = config.DEFAULT_CIRCULATION_POLICY
        total_copies = config.DEFAULT_TOTAL_COPIES
        language = config.DEFAULT_LANGUAGE

        return [
            isbn, title, author_str, publisher, pubdate, price,
            pages, binding, category, summary, cover_url, rating_average,
            location, status, availability, circulation_policy, total_copies, language
        ]

    def _save_failed_records(self):
        """保存失败的记录"""
        try:
            with open(config.FAILED_FILE, 'w', encoding='utf-8') as f:
                json.dump(self.failed_records, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"保存失败记录时出错: {e}")


def main():
    """主函数"""
    exporter = ExcelExporter()
    exporter.export_all()


if __name__ == "__main__":
    main()
